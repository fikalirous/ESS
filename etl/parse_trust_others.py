"""
ETL for the 'Trust in Others' sheet of ESS11_graphs_Ausra (version2)_20241209.xlsx

Layout discovered by manual inspection (openpyxl, data_only=True):
  - Rows 4-15:  Trend table. Row4 = headers (ESS round | Year | Average Trust in
    Other People (anweight)). Rows 5-15 = data for ESS rounds 1-11 (2002-2023).
  - Rows 20-24: ESS11 (2023) breakdown of trust by education. Row20 = 'ESS11'
    label, row21 = education-level column headers (B:F), rows 22-24 = the three
    trust bands (Lower Trust 0-4 / Neutral 5 / Higher Trust 6-10) with percentages
    per education level. This matches Figure 1.2 in the PDF report (tertiary =
    66% scoring 6-10).

Produces:
  - trust_others__trend.csv
  - trust_others__breakdown_education.csv
"""
import pathlib
import pandas as pd
import openpyxl

WORKBOOK = pathlib.Path(__file__).resolve().parent.parent / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "processed"
SHEET = "Trust in Others"

CATEGORY_MAP = {
    "Lower Trust (0-4)": "Lower Trust (0-4)",
    "Nautral answer (5)": "Neutral answer (5)",
    "Higher Trust (6-10)": "Higher Trust (6-10)",
}


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]

    # ---- Trend table: rows 5-15, cols A (round), B (year), C (value) ----
    trend_rows = []
    for r in range(5, 16):
        essround = ws.cell(row=r, column=1).value
        year = ws.cell(row=r, column=2).value
        value = ws.cell(row=r, column=3).value
        if essround is None or value is None:
            continue
        trend_rows.append(
            {
                "ess_round": int(essround),
                "year": str(int(year)) if isinstance(year, (int, float)) else str(year),
                "series": "Trust in Others",
                "value": round(float(value), 4),
            }
        )
    trend_df = pd.DataFrame(trend_rows)

    # ---- Education breakdown: header row21 (B:F education labels), rows22-24 ----
    edu_labels = {}
    for col in range(2, 7):  # B..F
        val = ws.cell(row=21, column=col).value
        if val is not None:
            edu_labels[col] = val

    breakdown_rows = []
    for r in range(22, 25):
        raw_category = ws.cell(row=r, column=1).value
        if raw_category is None:
            continue
        category = CATEGORY_MAP.get(raw_category, raw_category)
        for col, edu in edu_labels.items():
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            breakdown_rows.append(
                {
                    "category": category,
                    "group_type": "education",
                    "group_value": edu,
                    "value": round(float(val), 4),
                    "unit": "percent",
                }
            )
    breakdown_df = pd.DataFrame(breakdown_rows)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    trend_path = OUT_DIR / "trust_others__trend.csv"
    breakdown_path = OUT_DIR / "trust_others__breakdown_education.csv"
    trend_df.to_csv(trend_path, index=False)
    breakdown_df.to_csv(breakdown_path, index=False)

    print("=== trust_others__trend.csv ===")
    print(trend_df.head(11))
    print(f"rows: {len(trend_df)}  value range: {trend_df['value'].min():.2f} - {trend_df['value'].max():.2f}")
    print()
    print("=== trust_others__breakdown_education.csv ===")
    print(breakdown_df.head(15))
    print(f"rows: {len(breakdown_df)}  value range: {breakdown_df['value'].min():.2f} - {breakdown_df['value'].max():.2f}")
    print()
    print(f"Wrote: {trend_path}")
    print(f"Wrote: {breakdown_path}")


if __name__ == "__main__":
    main()
