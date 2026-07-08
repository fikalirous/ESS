"""
ETL for the 'Life Sat and Happiness' sheet of
ESS11_graphs_Ausra (version2)_20241209.xlsx

Layout discovered by manual inspection (openpyxl, data_only=True):
  - Rows 4-77 (cols A-O): a set of 11 repeating blocks (one per ESS round),
    each headed by 'essround = N' (col A / col I) followed by a Stata-style
    summary-statistics table (Variable/Obs/Weight/Mean/Std.dev/Min/Max) for
    stflife (life satisfaction) and happy (happiness). This is a re-derivation
    of the exact same round-by-round means found in the cleaner trend table
    below, so it is NOT re-emitted separately (redundant / would just
    duplicate the trend CSV under a clunkier schema).
  - Rows 4-16 (cols Q-T): a ready-made tidy trend table -- ESS round (R1..R11),
    Year, Life Satisfaction mean, Happiness mean, 2002-2023. This is the
    source used for the trend CSV (matches report Figure 5.1: 2023 life
    satisfaction 7.38 (report rounds to 7.4), happiness 7.71 (report ~7.7)).
  - Rows 80-119: ESS11 (2023) 'Life Satisfaction and Employment Status' table
    (report Table 5.1). Row83 headers 'Percent Satisfied' / 'Percent
    Dissatisfied', rows 84-88 give the five employment-status groups (In paid
    work / Unemployed looking for job / Inactive not looking / Refusal-DK /
    Total Population) with the two percentages directly -- this matches the
    report table exactly and is used as-is. There are also more granular
    per-group Freq/Percent/Cum sub-tables further down (rows 89-119) that
    contain the identical percentages broken out with frequency counts; these
    are redundant with the compact summary table and are not re-emitted.

Produces:
  - life_satisfaction_happiness__trend.csv
  - life_satisfaction_happiness__breakdown_employment.csv
"""
import pathlib
import pandas as pd
import openpyxl

WORKBOOK = pathlib.Path(__file__).resolve().parent.parent / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "processed"
SHEET = "Life Sat and Happiness"


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- Trend table: rows 5-15, cols Q(round label) R(year) S(life sat) T(happy) ----
    trend_rows = []
    for r in range(5, 16):
        essround_label = ws.cell(row=r, column=17).value  # Q, e.g. 'R1'
        year = ws.cell(row=r, column=18).value  # R
        life_sat = ws.cell(row=r, column=19).value  # S
        happy = ws.cell(row=r, column=20).value  # T
        if essround_label is None:
            continue
        essround = int(str(essround_label).lstrip("R"))
        year_str = str(int(year)) if isinstance(year, (int, float)) else str(year)
        if life_sat is not None:
            trend_rows.append(
                {"ess_round": essround, "year": year_str, "series": "Life Satisfaction", "value": round(float(life_sat), 4)}
            )
        if happy is not None:
            trend_rows.append(
                {"ess_round": essround, "year": year_str, "series": "Happiness", "value": round(float(happy), 4)}
            )
    trend_df = pd.DataFrame(trend_rows)
    trend_path = OUT_DIR / "life_satisfaction_happiness__trend.csv"
    trend_df.to_csv(trend_path, index=False)

    # ---- Employment breakdown: rows 84-88, cols F(group) G(% satisfied) H(% dissatisfied) ----
    breakdown_rows = []
    for r in range(84, 89):
        group = ws.cell(row=r, column=6).value  # F
        pct_sat = ws.cell(row=r, column=7).value  # G
        pct_dissat = ws.cell(row=r, column=8).value  # H
        if group is None:
            continue
        if pct_sat is not None:
            breakdown_rows.append(
                {
                    "category": "Satisfied",
                    "group_type": "employment",
                    "group_value": group,
                    "value": round(float(pct_sat), 4),
                    "unit": "percent",
                }
            )
        if pct_dissat is not None:
            breakdown_rows.append(
                {
                    "category": "Dissatisfied",
                    "group_type": "employment",
                    "group_value": group,
                    "value": round(float(pct_dissat), 4),
                    "unit": "percent",
                }
            )
    breakdown_df = pd.DataFrame(breakdown_rows)
    breakdown_path = OUT_DIR / "life_satisfaction_happiness__breakdown_employment.csv"
    breakdown_df.to_csv(breakdown_path, index=False)

    print("=== life_satisfaction_happiness__trend.csv ===")
    print(trend_df.head(10))
    print(f"rows: {len(trend_df)}  value range: {trend_df['value'].min():.2f} - {trend_df['value'].max():.2f}")
    print()
    print("=== life_satisfaction_happiness__breakdown_employment.csv ===")
    print(breakdown_df)
    print(f"rows: {len(breakdown_df)}  value range: {breakdown_df['value'].min():.2f} - {breakdown_df['value'].max():.2f}")
    print()
    print(f"Wrote: {trend_path}")
    print(f"Wrote: {breakdown_path}")


if __name__ == "__main__":
    main()
