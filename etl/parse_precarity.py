"""
Parse the 'Precarity' sheet from the ESS11 Ireland report workbook into
tidy long-format CSVs.

This sheet is a Country-specific ("Michael") supplementary question - X5,
"My job is secure" - that was asked as a national/rotating item in
ESS rounds 2, 5 and 9 for Ireland (row 3 label "National Question ESS9";
rows 9-10/23-24 give the round->year mapping explicitly: R2=2004, R5=2010,
R9=2018). Row 6 gives the exact question wording:
  "X5. Using this card, please tell me how true the following statements
  is about your current job: My job is secure"

Two overlapping representations of the same underlying trend are present:
  Table 1 (rows 8-19, cols A/C/D/E/F "COUNTRY-SPECIFIC: MICHEAL (FREQ)" /
  "(%)"): raw weighted frequencies per round, plus an adjacent "%" block
  that is DEMONSTRABLY BROKEN - e.g. the "Not Aplicable" row's round-9 "%"
  value is 101.4, and the "Total" row's own "%" values (247/336/203) don't
  sum to 100 either. This block was excluded from the clean CSVs; only the
  raw frequency numbers from this table are kept (as a separate, clearly
  labelled "freq" trend file), since those look internally consistent
  (each round's 4 substantive-category + Not Applicable + Refusal + No
  Answer frequencies sum to that round's Total row).

  Table 2 (rows 22-31, "...of those who responded"): a cleanly recomputed
  percentage trend restricted to the 4 substantive response categories
  (Not at all / A little / Quite / Very true), which correctly sum to
  ~100% per round, plus an "Insecure"/"Secure" 2-way aggregation. This is
  used as the primary (percent) trend CSV.

Further down the sheet (rows 48-63, "Calculating Income quintile values
for 2018") is a scratch/working decile-level breakdown of the same
question for round 9 (2018) only, explicitly labelled by its own author as
provisional ("Convert to % (and double check woith Monika results)") and
containing a visible formula error (`#VALUE!` in the "Total (Valid Cases)"
row, and a "Total (Valid Cases)" that is far short of the round-9 N).
This decile table was EXCLUDED from the clean output for that reason.

Immediately below it (rows 66-73, "Now quintiles") is a cleaner,
apparently-final income-QUINTILE breakdown of the same round-9 data (5
groups: Bottom 20% / 2 / 3 / 4 / Top 20%), with percentages that correctly
sum to 1.0 per quintile and an Insecure/Secure roll-up already computed.
This is used as the breakdown-by-income CSV (round 9 / 2018 only - the
quintile cross-tab isn't available for any other round in this sheet).
"""
import os
import pandas as pd
import openpyxl

SRC = r"C:\Users\Gokul\OneDrive\Internship\ESS\Project Florish\ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = r"C:\Users\Gokul\OneDrive\Internship\ESS\Project Florish\data\processed"
SHEET = "Precarity"

ROUND_YEAR = {2: "2004", 5: "2010", 9: "2018"}


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[SHEET]

    # --- Trend 1: raw weighted frequency counts by round (Table 1, rows 11-17) ---
    freq_categories = [
        (11, "Not at all true"),
        (12, "A little true"),
        (13, "Quite true"),
        (14, "Very true"),
        (15, "Not Applicable"),
        (16, "Refusal/Don't know"),
        (17, "No Answer"),
    ]
    freq_rows = []
    for row, cat in freq_categories:
        for col, rnd in zip("CDE", (2, 5, 9)):
            val = ws[f"{col}{row}"].value
            freq_rows.append((rnd, ROUND_YEAR[rnd], cat, round(float(val), 4)))
    freq_df = pd.DataFrame(freq_rows, columns=["ess_round", "year", "series", "value"])
    freq_out = os.path.join(OUT_DIR, "precarity__trend_job_security_freq.csv")
    freq_df.to_csv(freq_out, index=False)

    # --- Trend 2: clean percent-of-respondents trend by round (Table 2, rows 25-31) ---
    pct_categories = [
        (25, "Not at all true"),
        (26, "A little true"),
        (27, "Quite true"),
        (28, "Very true"),
        (30, "Insecure (combined: not at all/a little true)"),
        (31, "Secure (combined: quite/very true)"),
    ]
    pct_rows = []
    for row, cat in pct_categories:
        for col, rnd in zip("BCD", (2, 5, 9)):
            val = ws[f"{col}{row}"].value
            pct_rows.append((rnd, ROUND_YEAR[rnd], cat, round(float(val), 4)))
    pct_df = pd.DataFrame(pct_rows, columns=["ess_round", "year", "series", "value"])
    pct_out = os.path.join(OUT_DIR, "precarity__trend_job_security_percent.csv")
    pct_df.to_csv(pct_out, index=False)

    # --- Breakdown: income quintile x job security, round 9 (2018) only (rows 68-72) ---
    quintile_rows_meta = [
        (68, "Bottom 20%"),
        (69, "2nd quintile"),
        (70, "3rd quintile"),
        (71, "4th quintile"),
        (72, "Top 20%"),
    ]
    # L,M,N,O = fraction (0-1) for Not at all/A little/Quite/Very true; T,U = fraction Insecure/Secure
    cat_cols = [("L", "Not at all true"), ("M", "A little true"), ("N", "Quite true"), ("O", "Very true"),
                ("T", "Insecure (combined: not at all/a little true)"), ("U", "Secure (combined: quite/very true)")]
    bd_rows = []
    for row, group_label in quintile_rows_meta:
        for col, cat in cat_cols:
            val = ws[f"{col}{row}"].value
            if val is None:
                continue
            bd_rows.append((cat, "income_quintile", group_label, round(float(val) * 100, 4), "percent"))
    bd_df = pd.DataFrame(bd_rows, columns=["category", "group_type", "group_value", "value", "unit"])
    bd_out = os.path.join(OUT_DIR, "precarity__breakdown_income_quintile_r9_2018.csv")
    bd_df.to_csv(bd_out, index=False)

    print("=== Precarity ETL summary ===")
    for path, df in [(freq_out, freq_df), (pct_out, pct_df), (bd_out, bd_df)]:
        print(f"\n{path}  ({len(df)} rows)")
        print(df.head(10).to_string(index=False))
        print(f"  value range: {df['value'].min()} - {df['value'].max()}")


if __name__ == "__main__":
    main()
