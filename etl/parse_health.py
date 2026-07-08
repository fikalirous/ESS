"""
ETL parser for the 'Health and Health Services' sheet of the ESS11 Ireland
workbook (ESS11_graphs_Ausra (version2)_20241209.xlsx).

The sheet is a printed-report style layout: it mixes
  - a single ESS-round time-trend table (cols F:Q, top of sheet)
  - eleven repeating 'essround = N' Freq/Percent/Cum distribution blocks
  - several cross-tab breakdown sections (employment, education, income
    decile, age) each built from a compact "summary" table (used for the
    chart) PLUS a set of detailed per-group Freq/Percent/Cum sub-tables.

Rather than hand-transcribing row numbers, this script locates every
Freq/Percent/Cum sub-table generically by scanning for the literal header
value 'Freq.', then reads back to find the group label it belongs to
(e.g. essround marker, employment status, education level, income decile,
age band). Blocks are classified into sections positionally (the sheet's
sections always appear in the same order), which was verified by manual
inspection of the raw sheet (see NOTES.md).

Output: tidy CSVs in data/processed/, prefixed "health__".
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
SRC = BASE / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = BASE / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SHEET = "Health and Health Services"


def r4(x):
    return round(float(x), 4)


def load_ws():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    return wb[SHEET]


def find_freq_headers(ws):
    """Return sorted list of (row, col) for every cell literally == 'Freq.'"""
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Freq.":
                hits.append((cell.row, cell.column))
    hits.sort()
    return hits


def group_label_above(ws, header_row, label_col, max_back=6):
    """Scan upward from header_row-1 for the nearest row with a non-empty
    value in columns 1..label_col; return the leftmost non-None value."""
    for rr in range(header_row - 1, max(header_row - max_back, 0), -1):
        for cc in range(1, label_col + 1):
            v = ws.cell(row=rr, column=cc).value
            if v is not None:
                return v
    return None


def read_block(ws, header_row, freq_col, end_row):
    """Read data rows below a Freq./Percent./Cum. header until end_row
    (exclusive), keeping only rows where both the label and freq cell are
    populated (skips spacer rows and the next block's group-label rows)."""
    label_col = freq_col - 1
    percent_col = freq_col + 1
    cum_col = freq_col + 2
    out = []
    for r in range(header_row + 1, end_row):
        lbl = ws.cell(row=r, column=label_col).value
        freq = ws.cell(row=r, column=freq_col).value
        if lbl is None or freq is None:
            continue
        if not isinstance(freq, (int, float)):
            continue
        if str(lbl).strip() == "Total":
            continue
        pct = ws.cell(row=r, column=percent_col).value
        cum = ws.cell(row=r, column=cum_col).value
        out.append({"category": str(lbl).strip(), "freq": freq, "percent": pct, "cum": cum})
    return out


def norm_decile(v):
    """Normalise an income-decile group label (numeric 1.0-10.0 or the
    text 'Bottom'/'Top') to the sheet's own Bottom/2/.../Top convention."""
    if isinstance(v, (int, float)):
        n = int(v)
        if n == 1:
            return "Bottom"
        if n == 10:
            return "Top"
        return str(n)
    return str(v).strip()


def norm_age(v):
    return re.sub(r"\s+", " ", str(v).strip())


def main():
    ws = load_ws()

    # ---- year / essround header row (row 3 has years, row 2 has round #) ----
    def fmt_year(yr):
        if isinstance(yr, float) and yr.is_integer():
            return str(int(yr))
        return str(yr)

    year_by_round = {}
    for c in range(7, 18):  # G..Q
        rnd = ws.cell(2, c).value
        yr = ws.cell(3, c).value
        if rnd is not None:
            year_by_round[int(rnd)] = fmt_year(yr)

    # =========================================================
    # 1) Trend table: self-reported health, ESS rounds 1-11 (Figure 3.1)
    # =========================================================
    trend_rows = []
    series_map = {4: "Very good health", 5: "Fair/bad/very bad health"}
    for r, series in series_map.items():
        for c in range(7, 18):
            rnd = ws.cell(2, c).value
            val = ws.cell(r, c).value
            if rnd is None or val is None:
                continue
            rnd = int(rnd)
            trend_rows.append(
                {
                    "ess_round": rnd,
                    "year": year_by_round.get(rnd, ""),
                    "series": series,
                    "value": r4(val),
                }
            )
    write_csv(
        trend_rows,
        ["ess_round", "year", "series", "value"],
        "health__trend_selfreported_health.csv",
    )

    # =========================================================
    # Locate every Freq./Percent./Cum. sub-table in the sheet
    # =========================================================
    headers = find_freq_headers(ws)  # 62 blocks total, in sheet order

    # Positional classification of the 62 blocks (verified manually):
    #   0-10   essround = 1..11        (health_cat: Very good/Good/Fair-bad-verybad)
    #   11-12  employment status       (health_cat) - In paid work, Unemployed
    #   13-17  education level         (health_cat) - 5 groups
    #   18-27  income decile #1        (good_health) - 1.0..10.0  [authoritative]
    #   28-34  age x healthcare sat.   (sat_health) - 7 age bands
    #   35-41  age x self-rated health (health_cat) - 7 age bands
    #   42-51  income decile #2        (good_health dup, minor value drift - SKIPPED, see NOTES)
    #   52-61  income decile x happy   (happy_b) - 10 deciles
    assert len(headers) == 62, f"expected 62 Freq. blocks, found {len(headers)}"

    def block_bounds(i):
        row, col = headers[i]
        end = headers[i + 1][0] if i + 1 < len(headers) else ws.max_row + 1
        return row, col, end

    # ---- 2) Distribution: self-reported health by ESS round (0-10) ----
    dist_rows = []
    for i in range(0, 11):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = group_label_above(ws, header_row, label_col)
        m = re.search(r"essround\s*=\s*(\d+)", str(grp))
        rnd = int(m.group(1))
        for rec in read_block(ws, header_row, freq_col, end_row):
            dist_rows.append(
                {
                    "ess_round": rnd,
                    "year": year_by_round.get(rnd, ""),
                    "response_category": rec["category"],
                    "freq": r4(rec["freq"]),
                    "percent": r4(rec["percent"]),
                    "cum_percent": r4(rec["cum"]) if rec["cum"] is not None else "",
                }
            )
    write_csv(
        dist_rows,
        ["ess_round", "year", "response_category", "freq", "percent", "cum_percent"],
        "health__distribution_selfreported_health.csv",
    )

    # ---- 3) Breakdown: employment status (11-12) + education (13-17) ----
    empl_edu_rows = []
    for i in range(11, 13):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = str(group_label_above(ws, header_row, label_col)).strip()
        for rec in read_block(ws, header_row, freq_col, end_row):
            empl_edu_rows.append(
                {
                    "category": rec["category"],
                    "group_type": "employment",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    for i in range(13, 18):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = str(group_label_above(ws, header_row, label_col)).strip()
        for rec in read_block(ws, header_row, freq_col, end_row):
            empl_edu_rows.append(
                {
                    "category": rec["category"],
                    "group_type": "education",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    write_csv(
        empl_edu_rows,
        ["category", "group_type", "group_value", "value", "unit"],
        "health__breakdown_employment_education.csv",
    )

    # ---- 4) Breakdown: income decile - good health (18-27) + happy (52-61) ----
    RENAME_GOOD_HEALTH = {"Good": "Good health", "Otherwise": "Not good health"}
    decile_rows = []
    for i in range(18, 28):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = group_label_above(ws, header_row, label_col)
        grp = norm_decile(grp)
        for rec in read_block(ws, header_row, freq_col, end_row):
            cat = RENAME_GOOD_HEALTH.get(rec["category"], rec["category"])
            decile_rows.append(
                {
                    "category": cat,
                    "group_type": "income_decile",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    for i in range(52, 62):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = group_label_above(ws, header_row, label_col)
        grp = norm_decile(grp)
        for rec in read_block(ws, header_row, freq_col, end_row):
            decile_rows.append(
                {
                    "category": rec["category"],  # Happy / Unhappy
                    "group_type": "income_decile",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    write_csv(
        decile_rows,
        ["category", "group_type", "group_value", "value", "unit"],
        "health__breakdown_income_decile.csv",
    )

    # ---- 5) Breakdown: age x healthcare satisfaction (28-34) + age x self-rated health (35-41) ----
    RENAME_SAT = {
        "Satisfied": "Satisfied (healthcare system)",
        "Dissatisfied": "Dissatisfied (healthcare system)",
    }
    RENAME_HEALTH = {
        "Very good": "Very good (self-rated health)",
        "Good": "Good (self-rated health)",
        "Fair/bad/very bad": "Fair/bad/very bad (self-rated health)",
    }
    age_rows = []
    for i in range(28, 35):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = norm_age(group_label_above(ws, header_row, label_col))
        for rec in read_block(ws, header_row, freq_col, end_row):
            cat = RENAME_SAT.get(rec["category"], rec["category"])
            age_rows.append(
                {
                    "category": cat,
                    "group_type": "age",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    for i in range(35, 42):
        header_row, freq_col, end_row = block_bounds(i)
        label_col = freq_col - 1
        grp = norm_age(group_label_above(ws, header_row, label_col))
        for rec in read_block(ws, header_row, freq_col, end_row):
            cat = RENAME_HEALTH.get(rec["category"], rec["category"])
            age_rows.append(
                {
                    "category": cat,
                    "group_type": "age",
                    "group_value": grp,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    write_csv(
        age_rows,
        ["category", "group_type", "group_value", "value", "unit"],
        "health__breakdown_age.csv",
    )

    print("\nDone. health__*.csv files written to", OUT_DIR)


def write_csv(rows, fieldnames, filename):
    import csv

    path = OUT_DIR / filename
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"Wrote {len(rows):4d} rows -> {path.name}")

    # quick sanity print
    import pandas as pd

    df = pd.read_csv(path)
    print(df.head(3).to_string(index=False))
    if "value" in df.columns:
        print(f"  value range: {df['value'].min()} - {df['value'].max()}")
    if "percent" in df.columns:
        print(f"  percent range: {df['percent'].min()} - {df['percent'].max()}")
    print()


if __name__ == "__main__":
    main()
