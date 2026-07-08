"""
ETL for the 'Climate Change' sheet of ESS11_graphs_Ausra (version2)_20241209.xlsx

Layout discovered by manual inspection (openpyxl, data_only=True): a small
wide table, rows 4-8.
  - Row4: ESS round numbers in cols B/C/D (9, 10, 11)
  - Row5: corresponding years in cols B/C/D (2016, '2021/22', 2023)
  - Rows 6-8: one row per climate-perception question (col A = question text),
    mean scores (0-10 scale) per round in cols B/C/D.

Matches report Figure 9.1 (2023 means: reduce-climate-change likelihood
~6.11, people-will-limit-energy ~4.48/4.5, governments-will-act ~5.0).

Produces:
  - climate_change__trend.csv
"""
import pathlib
import pandas as pd
import openpyxl

WORKBOOK = pathlib.Path(__file__).resolve().parent.parent / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "processed"
SHEET = "Climate Change"

# shorten the (very long) verbatim question text into stable series labels
SERIES_LABELS = {
    6: "Limiting energy use would reduce climate change",
    7: "People will actually limit energy use",
    8: "Governments will take sufficient action",
}


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    round_cols = {2: ws.cell(row=4, column=2).value, 3: ws.cell(row=4, column=3).value, 4: ws.cell(row=4, column=4).value}
    year_cols = {2: ws.cell(row=5, column=2).value, 3: ws.cell(row=5, column=3).value, 4: ws.cell(row=5, column=4).value}

    rows = []
    for row_idx, series in SERIES_LABELS.items():
        for col, essround in round_cols.items():
            val = ws.cell(row=row_idx, column=col).value
            if val is None or essround is None:
                continue
            year = year_cols[col]
            year_str = str(int(year)) if isinstance(year, (int, float)) else str(year)
            rows.append(
                {
                    "ess_round": int(essround),
                    "year": year_str,
                    "series": series,
                    "value": round(float(val), 4),
                }
            )
    trend_df = pd.DataFrame(rows).sort_values(["series", "ess_round"]).reset_index(drop=True)
    trend_path = OUT_DIR / "climate_change__trend.csv"
    trend_df.to_csv(trend_path, index=False)

    print("=== climate_change__trend.csv ===")
    print(trend_df)
    print(f"rows: {len(trend_df)}  value range: {trend_df['value'].min():.2f} - {trend_df['value'].max():.2f}")
    print()
    print(f"Wrote: {trend_path}")


if __name__ == "__main__":
    main()
