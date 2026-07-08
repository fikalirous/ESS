"""
ETL parser for the 'Relig' sheet of
ESS11_graphs_Ausra (version2)_20241209.xlsx

NOTE: the separate 'Religion' tab is an older superseded draft (dated 2020, only goes to
ESS round 9) and is intentionally NOT read by this script.

Produces tidy CSVs in data/processed/:
  - religion__trend.csv                (ESS round 1-11, % belonging to a religion, "Yes")
  - religion__distribution.csv         (per-round Freq/Percent/Cum for Yes/No belonging)
  - religion__breakdown_gender.csv     (Table 8.2: % Yes/No by gender)
  - religion__breakdown_age.csv        (Table 8.2: % Yes/No by age group)
  - religion__breakdown_education.csv  (Table 8.2: % Yes/No by education)
  - religion__breakdown_urban_rural.csv (Table 8.2: % Yes/No by urban/rural location)

Judgment call: the sheet also contains a "Born in country" x "Belonging" cross-tab count
trend (cols I-U, rows ~29-49) which is out of scope for the requested breakdowns
(gender/age/education/location) and is intentionally not extracted -- see NOTES.md.

Run standalone: python parse_religion.py
"""
import os
import openpyxl
import pandas as pd

BASE_DIR = r"C:\Users\Gokul\OneDrive\Internship\ESS\Project Florish"
XLSX_PATH = os.path.join(BASE_DIR, "ESS11_graphs_Ausra (version2)_20241209.xlsx")
OUT_DIR = os.path.join(BASE_DIR, "data", "processed")
SHEET_NAME = "Relig"

os.makedirs(OUT_DIR, exist_ok=True)


def load_ws():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    return wb[SHEET_NAME]


def cell(ws, r, c):
    return ws.cell(row=r, column=c).value


def clean_num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(v), 4)
    except (ValueError, TypeError):
        return None


def year_to_str(v):
    if v is None or v == "":
        return None
    if isinstance(v, str):
        return v.strip()
    return str(int(v))


# ---------------------------------------------------------------------------
# Trend: row4 = ESS round (cols H..R), row5 = year, row6 = % Yes
# ---------------------------------------------------------------------------
def parse_trend(ws):
    round_row, year_row, pct_row = 4, 5, 6
    rows = []
    for c in range(8, 19):  # H(8) .. R(18)
        ess_round = cell(ws, round_row, c)
        if ess_round is None:
            continue
        ess_round = int(ess_round)
        year_s = year_to_str(cell(ws, year_row, c))
        val = clean_num(cell(ws, pct_row, c))
        if val is None:
            continue
        rows.append({
            "ess_round": ess_round, "year": year_s,
            "series": "Belonging to a religion (% Yes)", "value": val,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Distribution: per-round Freq/Percent/Cum for Yes/No, found via "-> essround" anchors
# ---------------------------------------------------------------------------
def parse_distribution(ws, round_year_map):
    rows = []
    anchor_rows = []
    for r in range(1, ws.max_row + 1):
        v = cell(ws, r, 1)
        if v is not None and str(v).strip().startswith("-> essround"):
            round_num = cell(ws, r, 3)  # col C
            if round_num is not None:
                anchor_rows.append((r, int(round_num)))

    for idx, (anchor_r, round_num) in enumerate(anchor_rows):
        next_anchor_r = anchor_rows[idx + 1][0] if idx + 1 < len(anchor_rows) else ws.max_row + 1
        search_limit = min(next_anchor_r, anchor_r + 30)
        # find header row containing 'Freq.' in cols B-F
        freq_col = None
        header_row = None
        for r in range(anchor_r, search_limit):
            for c in range(2, 7):
                v = cell(ws, r, c)
                if v is not None and str(v).strip() == "Freq.":
                    freq_col = c
                    header_row = r
                    break
            if freq_col:
                break
        if freq_col is None:
            continue
        pct_col = freq_col + 1
        cum_col = freq_col + 2
        year_s = round_year_map.get(round_num)
        for r in range(header_row, search_limit):
            label = cell(ws, r, 1)
            if label is None:
                continue
            label_s = str(label).strip()
            if label_s in ("Yes", "No"):
                freq = clean_num(cell(ws, r, freq_col))
                pct = clean_num(cell(ws, r, pct_col))
                cum = clean_num(cell(ws, r, cum_col))
                rows.append({
                    "ess_round": round_num, "year": year_s,
                    "response_category": label_s,
                    "freq": freq, "percent": pct, "cum_percent": cum,
                })
    return pd.DataFrame(rows)


def build_round_year_map(ws):
    d = {}
    for c in range(8, 19):
        rnd = cell(ws, 4, c)
        yr = cell(ws, 5, c)
        if rnd is not None:
            d[int(rnd)] = year_to_str(yr)
    return d


# ---------------------------------------------------------------------------
# Breakdown Table 8.2 (rows 162-185): cols G(label) H(%Yes) I(%No)
# Sections: Gender(167-168), Age(170-176), Education(178-182), Location(184-185)
# ---------------------------------------------------------------------------
def parse_breakdown_table82(ws):
    sections = {
        "gender": range(167, 169),
        "age": range(170, 177),
        "education": range(178, 183),
        "urban_rural": range(184, 186),
    }
    dfs = {}
    for group_type, row_range in sections.items():
        rows = []
        for r in row_range:
            label = cell(ws, r, 7)  # G
            if label is None:
                continue
            label_s = str(label).strip()
            pct_yes = clean_num(cell(ws, r, 8))  # H
            pct_no = clean_num(cell(ws, r, 9))  # I
            if pct_yes is not None:
                rows.append({
                    "category": "Yes", "group_type": group_type,
                    "group_value": label_s, "value": pct_yes, "unit": "percent",
                })
            if pct_no is not None:
                rows.append({
                    "category": "No", "group_type": group_type,
                    "group_value": label_s, "value": pct_no, "unit": "percent",
                })
        dfs[group_type] = pd.DataFrame(rows)
    return dfs


def main():
    ws = load_ws()
    round_year_map = build_round_year_map(ws)

    breakdowns = parse_breakdown_table82(ws)

    outputs = {
        "religion__trend.csv": parse_trend(ws),
        "religion__distribution.csv": parse_distribution(ws, round_year_map),
        "religion__breakdown_gender.csv": breakdowns["gender"],
        "religion__breakdown_age.csv": breakdowns["age"],
        "religion__breakdown_education.csv": breakdowns["education"],
        "religion__breakdown_urban_rural.csv": breakdowns["urban_rural"],
    }

    print(f"=== {SHEET_NAME} ===")
    for fname, df in outputs.items():
        out_path = os.path.join(OUT_DIR, fname)
        df.to_csv(out_path, index=False)
        print(f"\n--> {fname}: {len(df)} rows")
        print(df.head(8).to_string(index=False))
        if "value" in df.columns:
            print(f"    value range: [{df['value'].min()}, {df['value'].max()}]")
        if "percent" in df.columns:
            print(f"    percent range: [{df['percent'].min()}, {df['percent'].max()}]")

    print("\nDone. Files written to:", OUT_DIR)


if __name__ == "__main__":
    main()
