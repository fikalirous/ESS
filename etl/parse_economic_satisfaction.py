"""
ETL for the 'Econ Sat and HH Inc Adequacy' sheet of
ESS11_graphs_Ausra (version2)_20241209.xlsx

NOTE: the workbook also has a separate 'Economic Sat' tab, dated 2020 and only
covering rounds up to 9 -- an older superseded draft. It is intentionally
ignored per instructions.

Layout discovered by manual inspection (openpyxl, data_only=True):
  - Rows 3-14 (cols F-H): a ready-made tidy trend table -- ESS round (R1..R11),
    Year, % Satisfied with the present state of the economy, 2002-2023. Used
    directly for the trend CSV (report Figure 6.1).
  - Rows 3-82 (cols A-D): 11 repeating per-round distribution blocks (round
    number in col A, e.g. 'A4=1.0'), each a satecon Freq/Percent/Cum table
    with Satisfied/Dissatisfied rows + Total. Used for the distribution CSV,
    one row pair per round (22 rows total). Year per round taken from the
    trend table above (same round -> year mapping).
  - Rows 86-136: ESS11 (2023) satisfaction with the economy by age group
    (report Figure 6.2). Row88 has age-group column headers, but the cleanest
    source is the set of 7 per-age-group Freq/Percent/Cum sub-tables (each
    with a 'Satisfied' and 'Dissatisfied' row) starting at rows 91, 98, 105,
    112, 119, 126, 133 -- used for the breakdown CSV (both Satisfied and
    Dissatisfied percentages per age group, more complete than the single
    'Satisfied' summary row at row89).
  - Rows 138-212: ESS11 (2023) perception of household income adequacy by
    income quintile (report Figure 6.3). Rows 140-144 hold a compact
    ready-made matrix (quintile columns Bottom 20% / 2 / 3 / 4 / Top 20% / All
    in cols G-M, category rows 'Living comfortably' / 'Coping' / 'Finding it
    difficult' / 'Finding it very difficult' in col F) -- used directly for
    the breakdown CSV. The more verbose per-quintile Freq/Percent/Cum
    sub-tables further down (rows 144-212) contain the same percentages under
    slightly different category labels ('Living comfortably on present
    income' etc.) and are redundant, so not re-emitted separately.

Produces:
  - economic_satisfaction__trend.csv
  - economic_satisfaction__distribution.csv
  - economic_satisfaction__breakdown_age.csv
  - economic_satisfaction__breakdown_income_quintile.csv
"""
import re
import pathlib
import pandas as pd
import openpyxl

WORKBOOK = pathlib.Path(__file__).resolve().parent.parent / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "processed"
SHEET = "Econ Sat and HH Inc Adequacy"


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- Trend table: rows 4-14, cols F(round label) G(year) H(% satisfied) ----
    round_year = {}
    trend_rows = []
    for r in range(4, 15):
        round_label = ws.cell(row=r, column=6).value  # F, e.g. 'R1'
        year = ws.cell(row=r, column=7).value  # G
        pct = ws.cell(row=r, column=8).value  # H
        if round_label is None:
            continue
        essround = int(re.sub(r"\D", "", str(round_label)))
        year_str = str(int(year)) if isinstance(year, (int, float)) else str(year)
        round_year[essround] = year_str
        if pct is not None:
            trend_rows.append(
                {"ess_round": essround, "year": year_str, "series": "Economic Satisfaction", "value": round(float(pct), 4)}
            )
    trend_df = pd.DataFrame(trend_rows)
    trend_path = OUT_DIR / "economic_satisfaction__trend.csv"
    trend_df.to_csv(trend_path, index=False)

    # ---- Distribution tables: rows 3-82, per-round blocks in cols A-D ----
    dist_rows = []
    current_round = None
    for r in range(3, 83):
        col_a = ws.cell(row=r, column=1).value
        if isinstance(col_a, (int, float)) and col_a == int(col_a) and 1 <= col_a <= 11:
            # could be a round-number header OR a Satisfied/Dissatisfied freq value in col B... but
            # round headers are floats in col A on their own row with no sibling B/C/D values
            b_val = ws.cell(row=r, column=2).value
            if b_val is None:
                current_round = int(col_a)
                continue
        if col_a in ("Satisfied", "Dissatisfied") and current_round is not None:
            freq = ws.cell(row=r, column=2).value
            pct = ws.cell(row=r, column=3).value
            cum = ws.cell(row=r, column=4).value
            dist_rows.append(
                {
                    "ess_round": current_round,
                    "year": round_year.get(current_round, ""),
                    "response_category": col_a,
                    "freq": round(float(freq), 4) if freq is not None else None,
                    "percent": round(float(pct), 4) if pct is not None else None,
                    "cum_percent": round(float(cum), 4) if cum is not None else None,
                }
            )
    dist_df = pd.DataFrame(dist_rows)
    dist_path = OUT_DIR / "economic_satisfaction__distribution.csv"
    dist_df.to_csv(dist_path, index=False)

    # ---- Age breakdown: 7 per-age-group sub-tables, rows ~88-136 ----
    age_breakdown_rows = []
    current_age_group = None
    for r in range(88, 137):
        col_a = ws.cell(row=r, column=1).value
        if isinstance(col_a, str) and col_a not in ("satecon", "Satisfied", "Dissatisfied", "Total"):
            # a new age-group header, e.g. '25 and under', '26-35', ...
            current_age_group = col_a.strip()
            continue
        if col_a in ("Satisfied", "Dissatisfied") and current_age_group is not None:
            pct = ws.cell(row=r, column=3).value  # C
            if pct is not None:
                age_breakdown_rows.append(
                    {
                        "category": col_a,
                        "group_type": "age",
                        "group_value": current_age_group,
                        "value": round(float(pct), 4),
                        "unit": "percent",
                    }
                )
    age_df = pd.DataFrame(age_breakdown_rows)
    age_path = OUT_DIR / "economic_satisfaction__breakdown_age.csv"
    age_df.to_csv(age_path, index=False)

    # ---- Income quintile breakdown: matrix rows 140-144 ----
    quintile_cols = {
        7: "Bottom 20%",  # G
        8: "2nd quintile",  # H
        9: "3rd quintile",  # I
        10: "4th quintile",  # J
        11: "Top 20%",  # K
        13: "All",  # M
    }
    income_rows = []
    for r in range(141, 145):
        category = ws.cell(row=r, column=6).value  # F
        if category is None:
            continue
        for col, quintile in quintile_cols.items():
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            income_rows.append(
                {
                    "category": category,
                    "group_type": "income_quintile",
                    "group_value": quintile,
                    "value": round(float(val), 4),
                    "unit": "percent",
                }
            )
    income_df = pd.DataFrame(income_rows)
    income_path = OUT_DIR / "economic_satisfaction__breakdown_income_quintile.csv"
    income_df.to_csv(income_path, index=False)

    # ---- Summary / sanity output ----
    print("=== economic_satisfaction__trend.csv ===")
    print(trend_df.head(11))
    print(f"rows: {len(trend_df)}  value range: {trend_df['value'].min():.2f} - {trend_df['value'].max():.2f}")
    print()
    print("=== economic_satisfaction__distribution.csv ===")
    print(dist_df.head(10))
    print(f"rows: {len(dist_df)}  percent range: {dist_df['percent'].min():.2f} - {dist_df['percent'].max():.2f}")
    print()
    print("=== economic_satisfaction__breakdown_age.csv ===")
    print(age_df)
    print(f"rows: {len(age_df)}  value range: {age_df['value'].min():.2f} - {age_df['value'].max():.2f}")
    print()
    print("=== economic_satisfaction__breakdown_income_quintile.csv ===")
    print(income_df)
    print(f"rows: {len(income_df)}  value range: {income_df['value'].min():.2f} - {income_df['value'].max():.2f}")
    print()
    print(f"Wrote: {trend_path}")
    print(f"Wrote: {dist_path}")
    print(f"Wrote: {age_path}")
    print(f"Wrote: {income_path}")


if __name__ == "__main__":
    main()
