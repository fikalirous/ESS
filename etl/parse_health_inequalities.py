"""
ETL parser for the 'Hlth inequalities' sheet of the ESS11 Ireland workbook
(ESS11_graphs_Ausra (version2)_20241209.xlsx).

This sheet is the rotating-module "Social Inequalities in Health" data for
ESS Round 11 (2023) only - there is no multi-round trend here. It contains
four logical sections, each built from a compact "summary" table used for
the report chart PLUS a set of detailed checkbox (Marked/Not marked) or
Likert Freq/Percent/Cum sub-tables:

  1. Health problems in the last 12 months (12 checkbox items)
  2. Unable to get medical consultation/treatment in the last 12 months (Yes/No)
  3. Feelings in the past week (8-item CES-D style battery, 4-point frequency scale)
  4. Problems with accommodation (8 checkbox items)

A quirk of this workbook: long row labels in column A were typed as
several stacked single-word/phrase fragments across consecutive rows
(word-wrap done by hand, not real Excel merged cells), so the label text
cannot be reliably reconstructed by string concatenation alone. Instead,
each detailed checkbox sub-table's "Marked" percent is matched by VALUE
against the clean item names in the section's own summary chart table
(all percents within a section are unique to 2dp), which is a robust way
to recover the item name. See NOTES.md for details.
"""
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent.parent
SRC = BASE / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = BASE / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SHEET = "Hlth inequalities"
ESS_ROUND = 11
YEAR = "2023"


def r4(x):
    return round(float(x), 4)


def load_ws():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    return wb[SHEET]


def find_freq_headers(ws):
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Freq.":
                hits.append((cell.row, cell.column))
    hits.sort()
    return hits


def read_block(ws, header_row, freq_col, end_row):
    """Read Marked/Not marked (or similar) data rows below a Freq. header."""
    label_col = freq_col - 1
    percent_col = freq_col + 1
    cum_col = freq_col + 2
    out = []
    for r in range(header_row + 1, end_row):
        lbl = ws.cell(row=r, column=label_col).value
        freq = ws.cell(row=r, column=freq_col).value
        if lbl is None or freq is None:
            continue
        if not isinstance(freq, (int, float)):
            continue
        if str(lbl).strip() == "Total":
            continue
        pct = ws.cell(row=r, column=percent_col).value
        cum = ws.cell(row=r, column=cum_col).value
        out.append({"category": str(lbl).strip(), "freq": freq, "percent": pct, "cum": cum})
    return out


def write_csv(rows, fieldnames, filename):
    import csv

    path = OUT_DIR / filename
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    print(f"Wrote {len(rows):4d} rows -> {path.name}")

    import pandas as pd

    df = pd.read_csv(path)
    print(df.head(3).to_string(index=False))
    if "value" in df.columns:
        print(f"  value range: {df['value'].min()} - {df['value'].max()}")
    if "percent" in df.columns:
        print(f"  percent range: {df['percent'].min()} - {df['percent'].max()}")
    print()


def main():
    ws = load_ws()
    headers = find_freq_headers(ws)
    assert len(headers) == 29, f"expected 29 Freq. blocks, found {len(headers)}"

    def block_bounds(i):
        row, col = headers[i]
        end = headers[i + 1][0] if i + 1 < len(headers) else ws.max_row + 1
        return row, col, end

    # =========================================================
    # Section 1: Health problems, last 12 months (blocks 0-11)
    # Summary chart table lives at F6:G17 - 12 items with clean names.
    # =========================================================
    summary_health_problems = {}
    for r in range(6, 18):
        name = ws.cell(r, 6).value
        pct = ws.cell(r, 7).value
        if name is not None and pct is not None:
            summary_health_problems[round(float(pct), 2)] = str(name).strip()

    hp_rows = []
    for i in range(0, 12):
        header_row, freq_col, end_row = block_bounds(i)
        recs = read_block(ws, header_row, freq_col, end_row)
        marked = next(r for r in recs if r["category"] == "Marked")
        pct_key = round(float(marked["percent"]), 2)
        item_name = summary_health_problems.get(pct_key, f"UNKNOWN (row {header_row})")
        hp_rows.append(
            {
                "ess_round": ESS_ROUND,
                "year": YEAR,
                "response_category": item_name,
                "freq": r4(marked["freq"]),
                "percent": r4(marked["percent"]),
                "cum_percent": "",  # independent checkbox items - see NOTES.md
            }
        )
    write_csv(
        hp_rows,
        ["ess_round", "year", "response_category", "freq", "percent", "cum_percent"],
        "health_inequalities__distribution_health_problems_12m.csv",
    )

    # =========================================================
    # Section 2: Unable to get medical consultation/treatment, last 12 months
    # (block 12) - simple Yes/No distribution.
    # =========================================================
    header_row, freq_col, end_row = block_bounds(12)
    recs = read_block(ws, header_row, freq_col, end_row)
    med_rows = []
    for rec in recs:
        med_rows.append(
            {
                "ess_round": ESS_ROUND,
                "year": YEAR,
                "response_category": rec["category"],
                "freq": r4(rec["freq"]),
                "percent": r4(rec["percent"]),
                "cum_percent": r4(rec["cum"]) if rec["cum"] is not None else "",
            }
        )
    write_csv(
        med_rows,
        ["ess_round", "year", "response_category", "freq", "percent", "cum_percent"],
        "health_inequalities__distribution_unable_medical_treatment.csv",
    )

    # =========================================================
    # Section 3: Feelings in the past week (blocks 13-20)
    # Primary source: clean summary matrix at F177:J184 (8 items x 4 freq cats).
    # Supplementary: the 8 detailed Freq/Percent/Cum sub-tables (same item
    # order as the summary matrix) - values differ slightly from the summary
    # (different weighting/base?), kept as a separate file - see NOTES.md.
    # =========================================================
    freq_categories = [
        ws.cell(176, c).value for c in range(7, 11)
    ]  # G176:J176 header labels
    feelings_summary_rows = []
    item_order = []
    for r in range(177, 185):
        item = ws.cell(r, 6).value
        if item is None:
            continue
        item = str(item).strip()
        item_order.append(item)
        for c, cat in zip(range(7, 11), freq_categories):
            val = ws.cell(r, c).value
            if val is None:
                continue
            feelings_summary_rows.append(
                {
                    "category": cat,
                    "group_type": "feeling_item",
                    "group_value": item,
                    "value": r4(val),
                    "unit": "percent",
                }
            )
    write_csv(
        feelings_summary_rows,
        ["category", "group_type", "group_value", "value", "unit"],
        "health_inequalities__breakdown_feelings_past_week.csv",
    )

    feelings_detailed_rows = []
    for idx, i in enumerate(range(13, 21)):
        header_row, freq_col, end_row = block_bounds(i)
        item = item_order[idx] if idx < len(item_order) else f"item_{idx}"
        for rec in read_block(ws, header_row, freq_col, end_row):
            feelings_detailed_rows.append(
                {
                    "category": rec["category"],
                    "group_type": "feeling_item",
                    "group_value": item,
                    "value": r4(rec["percent"]),
                    "unit": "percent",
                }
            )
    write_csv(
        feelings_detailed_rows,
        ["category", "group_type", "group_value", "value", "unit"],
        "health_inequalities__breakdown_feelings_past_week_detailed.csv",
    )

    # =========================================================
    # Section 4: Problems with accommodation (blocks 21-28)
    # Summary chart table lives at F258:G265 - 8 items with clean names.
    # =========================================================
    summary_accom = {}
    for r in range(258, 266):
        name = ws.cell(r, 6).value
        pct = ws.cell(r, 7).value
        if name is not None and pct is not None:
            summary_accom[round(float(pct), 2)] = str(name).strip()

    accom_rows = []
    for i in range(21, 29):
        header_row, freq_col, end_row = block_bounds(i)
        recs = read_block(ws, header_row, freq_col, end_row)
        marked = next(r for r in recs if r["category"] == "Marked")
        pct_key = round(float(marked["percent"]), 2)
        item_name = summary_accom.get(pct_key, f"UNKNOWN (row {header_row})")
        accom_rows.append(
            {
                "ess_round": ESS_ROUND,
                "year": YEAR,
                "response_category": item_name,
                "freq": r4(marked["freq"]),
                "percent": r4(marked["percent"]),
                "cum_percent": "",  # independent checkbox items - see NOTES.md
            }
        )
    write_csv(
        accom_rows,
        ["ess_round", "year", "response_category", "freq", "percent", "cum_percent"],
        "health_inequalities__distribution_accommodation_problems.csv",
    )

    print("\nDone. health_inequalities__*.csv files written to", OUT_DIR)


if __name__ == "__main__":
    main()
