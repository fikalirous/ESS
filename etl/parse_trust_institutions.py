"""
ETL for the 'Trust in Institutions' sheet of
ESS11_graphs_Ausra (version2)_20241209.xlsx

NOTE: the workbook also has a separate 'Sheet1' tab which is an older, redundant
cut of the same education-breakdown data -- it is intentionally ignored per
instructions.

Layout discovered by manual inspection (openpyxl, data_only=True):
  - Rows 4-47: five Freq/Percent/Cum distribution tables, one per institution
    (Garda rows 4-11, Legal System 13-20, Parliament 22-29, Political Parties
    31-38, Politicians 40-47). Each block has an ESS9 cut on the left (cols A-D,
    labelled 'ESS9' in B2, effectively 2018) and an ESS11 cut on the right
    (cols Q-T, labelled 'ESS11' in Q2/Q3, effectively 2023), with response
    categories Low / Neutral (5) / High.
    There is ALSO a compact "Low vs High only" summary table crammed into the
    same row range (cols F-H for ESS9, cols V-X for ESS11) that just restates
    the Low/High percentages for all five institutions in one place (this is
    the source for report Figure 2.1). It is fully redundant with the five
    per-institution distribution tables below it, so it is skipped.
  - Rows 50-61: trend table, mean trust score (0-10) 2002-2023 for Garda, Legal
    System, Parliament, Politicians, Political Parties (Political Parties has
    no value for round 1 / 2002 - not asked that round).
  - Rows 71-77: ESS11 (2023) % with high trust (6-10) by education level, for
    Police(=Garda)/Legal System/Parliament/Political Parties/Politicians. This
    is report Figure 2.3.

ESS round -> year mapping used for the distribution tables (not stated
explicitly inside this sheet's distribution blocks, taken from the sheet's own
trend table / other sheets in the workbook): round 9 = 2018, round 11 = 2023.

Produces:
  - trust_institutions__trend.csv
  - trust_institutions__breakdown_education.csv
  - trust_institutions__distribution_garda.csv
  - trust_institutions__distribution_legal_system.csv
  - trust_institutions__distribution_parliament.csv
  - trust_institutions__distribution_political_parties.csv
  - trust_institutions__distribution_politicians.csv
"""
import pathlib
import pandas as pd
import openpyxl

WORKBOOK = pathlib.Path(__file__).resolve().parent.parent / "ESS11_graphs_Ausra (version2)_20241209.xlsx"
OUT_DIR = pathlib.Path(__file__).resolve().parent.parent / "data" / "processed"
SHEET = "Trust in Institutions"

RESPONSE_MAP = {
    "Low": "Low Trust (0-4)",
    "Nautral answer (5)": "Neutral answer (5)",
    "High": "High Trust (6-10)",
}

# (institution label used in output, header_row, table_start_row, slug)
INSTITUTIONS = [
    ("Garda", 4, 5, "garda"),
    ("Legal System", 13, 14, "legal_system"),
    ("Parliament", 22, 23, "parliament"),
    ("Political Parties", 31, 32, "political_parties"),
    ("Politicians", 40, 41, "politicians"),
]


def parse_distribution_block(ws, data_start_row):
    """Parse a 3-row Low/Neutral/High block for both ESS9 (cols A-D) and
    ESS11 (cols Q-T), returning a tidy list of dict rows."""
    rows = []
    # data rows are at data_start_row+2, +3, +4 (Low, Neutral, High) based on
    # the fixed pattern observed (one blank row after the Freq/Percent/Cum
    # header before the Low row starts)
    for offset in range(0, 6):
        r = data_start_row + offset
        label_ess9 = ws.cell(row=r, column=1).value  # col A
        label_ess11 = ws.cell(row=r, column=17).value  # col Q
        if label_ess9 in RESPONSE_MAP:
            freq = ws.cell(row=r, column=2).value
            pct = ws.cell(row=r, column=3).value
            cum = ws.cell(row=r, column=4).value
            rows.append(
                {
                    "ess_round": 9,
                    "year": "2018",
                    "response_category": RESPONSE_MAP[label_ess9],
                    "freq": round(float(freq), 4) if freq is not None else None,
                    "percent": round(float(pct), 4) if pct is not None else None,
                    "cum_percent": round(float(cum), 4) if cum is not None else None,
                }
            )
        if label_ess11 in ("Lower Trust (0-4)", "Nautral answer (5)", "Higher Trust (6-11)"):
            norm = {
                "Lower Trust (0-4)": "Low Trust (0-4)",
                "Nautral answer (5)": "Neutral answer (5)",
                "Higher Trust (6-11)": "High Trust (6-10)",
            }[label_ess11]
            freq = ws.cell(row=r, column=18).value  # col R
            pct = ws.cell(row=r, column=19).value  # col S
            cum = ws.cell(row=r, column=20).value  # col T
            rows.append(
                {
                    "ess_round": 11,
                    "year": "2023",
                    "response_category": norm,
                    "freq": round(float(freq), 4) if freq is not None else None,
                    "percent": round(float(pct), 4) if pct is not None else None,
                    "cum_percent": round(float(cum), 4) if cum is not None else None,
                }
            )
    return rows


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- Trend table: rows 51-61, cols Q(round) R(year) S..W(institutions) ----
    trend_cols = {
        19: "Garda",  # S
        20: "Legal System",  # T
        21: "Parliament",  # U
        22: "Politicians",  # V
        23: "Political Parties",  # W
    }
    trend_rows = []
    for r in range(51, 62):
        essround = ws.cell(row=r, column=17).value  # Q
        year = ws.cell(row=r, column=18).value  # R
        if essround is None:
            continue
        for col, series in trend_cols.items():
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            trend_rows.append(
                {
                    "ess_round": int(essround),
                    "year": str(int(year)) if isinstance(year, (int, float)) else str(year),
                    "series": series,
                    "value": round(float(val), 4),
                }
            )
    trend_df = pd.DataFrame(trend_rows)
    trend_path = OUT_DIR / "trust_institutions__trend.csv"
    trend_df.to_csv(trend_path, index=False)

    # ---- Education breakdown: header row72 (R:V institutions), rows73-77 ----
    edu_inst_cols = {
        18: "Garda",  # R  (labelled 'Police' in sheet)
        19: "Legal System",  # S
        20: "Parliament",  # T
        21: "Political Parties",  # U
        22: "Politicians",  # V
    }
    breakdown_rows = []
    for r in range(73, 78):
        edu = ws.cell(row=r, column=17).value  # Q
        if edu is None:
            continue
        for col, inst in edu_inst_cols.items():
            val = ws.cell(row=r, column=col).value
            if val is None:
                continue
            breakdown_rows.append(
                {
                    "category": inst,
                    "group_type": "education",
                    "group_value": edu.strip() if isinstance(edu, str) else edu,
                    "value": round(float(val), 4),
                    "unit": "percent",
                }
            )
    breakdown_df = pd.DataFrame(breakdown_rows)
    breakdown_path = OUT_DIR / "trust_institutions__breakdown_education.csv"
    breakdown_df.to_csv(breakdown_path, index=False)

    # ---- Distribution tables, one CSV per institution ----
    dist_paths = []
    for label, header_row, data_start_row, slug in INSTITUTIONS:
        rows = parse_distribution_block(ws, data_start_row)
        df = pd.DataFrame(rows)
        path = OUT_DIR / f"trust_institutions__distribution_{slug}.csv"
        df.to_csv(path, index=False)
        dist_paths.append((label, path, df))

    # ---- Summary / sanity output ----
    print("=== trust_institutions__trend.csv ===")
    print(trend_df.head(10))
    print(f"rows: {len(trend_df)}  value range: {trend_df['value'].min():.2f} - {trend_df['value'].max():.2f}")
    print()
    print("=== trust_institutions__breakdown_education.csv ===")
    print(breakdown_df.head(10))
    print(f"rows: {len(breakdown_df)}  value range: {breakdown_df['value'].min():.2f} - {breakdown_df['value'].max():.2f}")
    print()
    for label, path, df in dist_paths:
        print(f"=== {path.name} ({label}) ===")
        print(df)
        print(f"rows: {len(df)}")
        print()

    print(f"Wrote: {trend_path}")
    print(f"Wrote: {breakdown_path}")
    for _, path, _ in dist_paths:
        print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
