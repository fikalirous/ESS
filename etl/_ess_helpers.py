"""
Shared, lightweight helpers for parsing the ESS11 Ireland "printed report style"
workbook sheets. Not a public API - just deduplicated logic reused across the
four C19 / Precarity parser scripts in this folder. Each parse_*.py script
still runs standalone (this module has no side effects on import).
"""
import openpyxl


def load_ws(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb[sheet_name]


def find_freq_anchors(ws):
    """Find every (row, col) where col holds literal 'Freq.' immediately
    followed by 'Percent' and 'Cum.' in the next two columns - these anchor
    a vertical Freq/Percent/Cum distribution table."""
    anchors = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip() == 'Freq.':
                v2 = ws.cell(row=r, column=c + 1).value
                v3 = ws.cell(row=r, column=c + 2).value
                if (isinstance(v2, str) and v2.strip() == 'Percent' and
                        isinstance(v3, str) and v3.strip() == 'Cum.'):
                    anchors.append((r, c))
    return sorted(anchors)


def find_captions(ws, prefix='Figure'):
    """Find every 'Figure X: ...' caption cell -> (row, col, text)."""
    caps = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().startswith(prefix):
                caps.append((r, c, v.strip()))
    return sorted(caps)


def nearest_caption_text(anchor_row, caps):
    best, best_dist = None, None
    for (r, c, text) in caps:
        d = abs(r - anchor_row)
        if best_dist is None or d < best_dist:
            best_dist, best = d, text
    return best


def extract_vertical_table(ws, header_row, freq_col, max_scan_row):
    """Extract category rows below a Freq/Percent/Cum header.
    Label = the leftmost non-null cell in columns [1 .. freq_col-1] of the row
    (handles the couple of tables in this workbook where the intended label
    spilled into a column other than the one immediately left of Freq.).
    Stops at a row whose label is literally 'Total'."""
    rows = []
    total_freq = None
    r = header_row + 1
    blank_streak = 0
    while r <= max_scan_row:
        freq_v = ws.cell(row=r, column=freq_col).value
        pct_v = ws.cell(row=r, column=freq_col + 1).value
        cum_v = ws.cell(row=r, column=freq_col + 2).value
        label_v = None
        for cc in range(1, freq_col):
            vv = ws.cell(row=r, column=cc).value
            if vv is not None and str(vv).strip() != '':
                label_v = vv
                break
        if label_v is None and freq_v is None:
            blank_streak += 1
            if blank_streak > 3:
                break
            r += 1
            continue
        blank_streak = 0
        label_str = str(label_v).strip() if label_v is not None else None
        if label_str is not None and label_str.lower() == 'total':
            total_freq = freq_v
            break
        if label_v is not None and isinstance(freq_v, (int, float)):
            rows.append((label_v, freq_v, pct_v, cum_v))
        r += 1
    return rows, total_freq


GROUP_NAMES_C19_EXPERIENCE = [
    'Yes, I tested positive for COVID-19',
    'Yes, I think I had COVID-19 but was not tested/did not test positive',
    'No, I have not had COVID-19',
]


def find_crosstab_anchors(ws, group_names):
    anchors = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column - 1):
            v1 = ws.cell(row=r, column=c).value
            v2 = ws.cell(row=r, column=c + 1).value
            v3 = ws.cell(row=r, column=c + 2).value
            if (isinstance(v1, str) and v1.strip() == group_names[0] and
                    isinstance(v2, str) and v2.strip() == group_names[1] and
                    isinstance(v3, str) and v3.strip() == group_names[2]):
                anchors.append((r, c))
    return sorted(anchors)


def extract_crosstab(ws, header_row, group_start_col, max_rows=13):
    label_col = group_start_col - 1
    rows = []
    r = header_row + 1
    blank_streak = 0
    count = 0
    while count < max_rows:
        label_v = ws.cell(row=r, column=label_col).value
        v1 = ws.cell(row=r, column=group_start_col).value
        v2 = ws.cell(row=r, column=group_start_col + 1).value
        v3 = ws.cell(row=r, column=group_start_col + 2).value
        if label_v is None and v1 is None and v2 is None and v3 is None:
            blank_streak += 1
            if blank_streak > 2:
                break
            r += 1
            continue
        blank_streak = 0
        if label_v is not None:
            rows.append((label_v, v1, v2, v3))
            count += 1
        r += 1
    return rows


def scale_label(label_v, endpoint0, endpoint10):
    """Normalize an 11-point-scale row label (0/10 endpoints are text,
    1-9 are bare floats like 1.0, 2.0, ...) into a clean response_category
    string, using full (non-truncated) endpoint text supplied by the caller."""
    if isinstance(label_v, (int, float)):
        return str(int(label_v))
    s = str(label_v).strip()
    # truncated / partial matches -> map to canonical endpoint text
    if s.lower().startswith(endpoint0[:15].lower()) or 'much more important to prioritise public' in s.lower():
        return endpoint0
    if s.lower().startswith(endpoint10[:15].lower()):
        return endpoint10
    return s
